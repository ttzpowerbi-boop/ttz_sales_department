"""
🛡️ ARMOR HAND - Облачный Mini App v7.0
Изменения:
  - Убрана вкладка "Импорт Excel"
  - Выгрузка предзаказа в Excel точно по шаблону акта ТТЗ
"""

import os
import json
import io
import re
from datetime import datetime, date
from flask import Flask, render_template_string, request, jsonify, send_file
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ── openpyxl для Excel ──────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)
session = requests.Session()
retry = Retry(total=3, backoff_factor=1, status_forcelist=[429,500,502,503,504])
session.mount("http://", HTTPAdapter(max_retries=retry))
session.mount("https://", HTTPAdapter(max_retries=retry))

# ── In-memory хранилище предзаказов ─────────────────────────────────────────
ORDERS: dict = {}          # { order_id: { id, number, created_at, items, comment } }
ORDER_COUNTER: list = [0]  # используем список чтобы мутировать из функции

def next_order_id():
    ORDER_COUNTER[0] += 1
    today = datetime.now().strftime("%Y%m%d")
    num = f"ПЗ-{today}-{ORDER_COUNTER[0]:04d}"
    return ORDER_COUNTER[0], num

# ============================================================================
# HTML
# ============================================================================

MINI_APP_HTML = r'''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ARMOR HAND</title>
<script src="https://telegram.org/js/telegram-web-app.js" async></script>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f2f5;color:#333;min-height:100vh;}

/* ── error ── */
#error-screen{position:fixed;top:0;left:0;width:100%;height:100%;background:linear-gradient(135deg,#667eea,#764ba2);display:none;align-items:center;justify-content:center;z-index:9999;padding:20px;}
.error-box{background:#fff;padding:40px 30px;border-radius:12px;max-width:400px;text-align:center;}
.error-box h2{color:#c62828;font-size:24px;margin-bottom:12px;}

/* ── layout ── */
.app{display:none;}
.container{max-width:600px;margin:0 auto;background:#fff;min-height:100vh;}
.header{background:linear-gradient(135deg,#1e3c72,#2a5298);color:#fff;padding:18px 20px;text-align:center;position:sticky;top:0;z-index:10;}
.header h1{font-size:22px;margin-bottom:3px;}
.header p{font-size:13px;opacity:.85;}

/* ── nav tabs ── */
.nav{display:flex;background:#1a3460;}
.nav-btn{flex:1;padding:11px 4px;background:none;border:none;color:rgba(255,255,255,.65);font-size:12px;cursor:pointer;border-bottom:3px solid transparent;transition:.2s;}
.nav-btn.active{color:#fff;border-bottom-color:#64b5f6;}

/* ── content ── */
.content{padding:16px;}
.page{display:none;}
.page.active{display:block;}

/* ── search ── */
.search-box{display:flex;gap:8px;}
.search-input{flex:1;padding:11px 12px;border:2px solid #ddd;border-radius:8px;font-size:14px;}
.search-btn{padding:11px 18px;background:#2a5298;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;white-space:nowrap;}
.products{display:flex;flex-direction:column;gap:8px;max-height:380px;overflow-y:auto;margin-top:12px;}
.product{padding:12px;border:2px solid #e0e0e0;border-radius:8px;cursor:pointer;transition:.15s;}
.product:hover{border-color:#2a5298;background:#f0f4ff;}
.product-name{font-weight:600;color:#1e3c72;font-size:13px;}

/* ── cart / preview tables ── */
table{width:100%;border-collapse:collapse;margin:12px 0;font-size:13px;}
th,td{border:1px solid #ddd;padding:8px 10px;text-align:left;}
th{background:#e3f2fd;font-weight:600;}
.action-btn{padding:6px 10px;border:none;border-radius:6px;cursor:pointer;}
.edit-btn{background:#2196f3;color:#fff;}
.remove-btn{background:#f44336;color:#fff;}

/* ── buttons ── */
.btn{padding:13px;border:none;border-radius:8px;cursor:pointer;font-weight:600;color:#fff;font-size:14px;}
.btn-primary{background:#4caf50;}
.btn-secondary{background:#757575;}
.btn-danger{background:#f44336;}
.btn-blue{background:#1565c0;}
.btn-excel{background:#217346;}
.buttons{display:flex;gap:8px;margin-top:16px;flex-wrap:wrap;}
.buttons .btn{flex:1;min-width:120px;}

/* ── message toast ── */
.message{padding:11px 14px;border-radius:8px;margin-bottom:12px;display:none;text-align:center;font-weight:600;font-size:14px;}
.message.success{background:#c8e6c9;color:#2e7d32;display:block;}
.message.error{background:#ffcdd2;color:#c62828;display:block;}
.message.info{background:#e3f2fd;color:#1565c0;display:block;}

/* ── order number banner ── */
.order-number-banner{background:linear-gradient(135deg,#1e3c72,#2a5298);color:#fff;padding:14px 18px;border-radius:10px;margin-bottom:14px;text-align:center;}
.order-number-banner .num{font-size:20px;font-weight:700;letter-spacing:1px;}
.order-number-banner .label{font-size:12px;opacity:.8;margin-top:3px;}

/* ── orders list ── */
.orders-filter{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap;align-items:center;}
.orders-filter input[type=date]{padding:8px 10px;border:2px solid #ddd;border-radius:8px;font-size:13px;flex:1;min-width:130px;}
.orders-filter label{font-size:12px;color:#555;white-space:nowrap;}
.filter-row{display:flex;gap:8px;align-items:center;flex:1;}
.order-card{border:2px solid #e0e0e0;border-radius:10px;padding:14px;margin-bottom:10px;cursor:pointer;transition:.15s;}
.order-card:hover{border-color:#2a5298;background:#f0f4ff;}
.order-card-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;}
.order-card-num{font-weight:700;color:#1e3c72;font-size:15px;}
.order-card-date{font-size:12px;color:#888;}
.order-card-count{font-size:13px;color:#555;}
.empty-state{text-align:center;padding:40px 20px;color:#999;}
.empty-state .icon{font-size:48px;margin-bottom:12px;}

/* ── order detail ── */
.order-detail-header{background:#f5f7fa;border-radius:10px;padding:14px;margin-bottom:14px;}
.order-detail-header .num{font-size:18px;font-weight:700;color:#1e3c72;}
.order-detail-header .meta{font-size:12px;color:#666;margin-top:4px;}

textarea{width:100%;height:70px;padding:10px;border:2px solid #ddd;border-radius:8px;font-size:14px;resize:none;}

/* ── import excel section ── */
</style>
</head>
<body>

<div id="error-screen">
  <div class="error-box">
    <h2>🔒 Доступ запрещён</h2>
    <p>Приложение работает <strong>только внутри Telegram</strong>.</p>
    <p style="margin-top:8px;">Откройте через: <strong>@TTZ_Sales_Department_bot</strong></p>
  </div>
</div>

<div class="app">
  <div class="container">
    <div class="header">
      <h1>🛡️ ARMOR HAND</h1>
      <p id="headerTitle">Форма предзаказа</p>
    </div>

    <!-- TAB NAV -->
    <div class="nav">
      <button class="nav-btn active" onclick="switchTab('search')">🔍 Поиск</button>
      <button class="nav-btn" onclick="switchTab('orders')">📋 Предзаказы</button>

    </div>

    <div class="content">
      <div id="message" class="message"></div>

      <!-- ══════════════ PAGE: SEARCH ══════════════ -->
      <div id="searchPage" class="page active">
        <div class="search-box">
          <input type="text" id="searchInput" class="search-input" placeholder="Размер, тип, ГОСТ..."
            onkeydown="if(event.key==='Enter') searchProducts()">
          <button class="search-btn" onclick="searchProducts()">Найти</button>
        </div>
        <div id="productsList" class="products" style="display:none;"></div>
      </div>

      <!-- ══════════════ PAGE: CART ══════════════ -->
      <div id="cartPage" class="page">
        <button class="btn btn-secondary" onclick="backToSearch()" style="margin-bottom:12px;">← Назад</button>
        <h3>🛒 Корзина</h3>
        <table id="cartTable">
          <thead><tr><th>Товар</th><th>Кол-во</th><th></th><th></th></tr></thead>
          <tbody id="cartBody"></tbody>
        </table>
        <div class="buttons">
          <button class="btn btn-danger" onclick="clearCart()">Очистить</button>
          <button class="btn btn-primary" onclick="showPreview()">Предпросмотр →</button>
        </div>
      </div>

      <!-- ══════════════ PAGE: PREVIEW ══════════════ -->
      <div id="previewPage" class="page">
        <button class="btn btn-secondary" onclick="showPage('cart')" style="margin-bottom:12px;">← В корзину</button>
        <h3>📄 Предварительный просмотр</h3>
        <table>
          <thead><tr><th>Товар</th><th style="text-align:right;">Кол-во</th></tr></thead>
          <tbody id="previewTable"></tbody>
        </table>
        <div style="margin-bottom:8px;font-size:13px;color:#555;">Позиций: <b id="previewCount">0</b></div>
        <div>
          <label style="font-size:13px;font-weight:600;display:block;margin-bottom:6px;">Комментарий:</label>
          <textarea id="orderComment" placeholder="(необязательно)"></textarea>
        </div>
        <div class="buttons">
          <button class="btn btn-secondary" onclick="showPage('cart')">Отмена</button>
          <button class="btn btn-primary" onclick="confirmAndSend()">✅ Оформить предзаказ</button>
        </div>
      </div>

      <!-- ══════════════ PAGE: SUCCESS ══════════════ -->
      <div id="successPage" class="page">
        <div class="order-number-banner">
          <div class="label">✅ Предзаказ создан!</div>
          <div class="num" id="successOrderNum">—</div>
          <div class="label" id="successOrderDate"></div>
        </div>
        <p style="font-size:14px;color:#555;margin-bottom:16px;text-align:center;">
          Консультант свяжется с вами в ближайшее время.
        </p>
        <div class="buttons">
          <button class="btn btn-blue" onclick="viewLastOrder()">👁 Открыть заказ</button>
          <button class="btn btn-excel" onclick="downloadLastOrder()">⬇ Скачать Excel</button>
        </div>
        <div class="buttons" style="margin-top:8px;">
          <button class="btn btn-secondary" onclick="newOrder()">+ Новый предзаказ</button>
        </div>
      </div>

      <!-- ══════════════ PAGE: ORDERS LIST ══════════════ -->
      <div id="ordersPage" class="page">
        <div class="orders-filter">
          <div class="filter-row">
            <label>С:</label>
            <input type="date" id="filterFrom">
            <label>По:</label>
            <input type="date" id="filterTo">
          </div>
          <button class="btn btn-blue" style="flex:0 0 auto;padding:8px 14px;font-size:13px;" onclick="renderOrdersList()">Фильтр</button>
        </div>
        <div id="ordersList"></div>
      </div>

      <!-- ══════════════ PAGE: ORDER DETAIL ══════════════ -->
      <div id="orderDetailPage" class="page">
        <button class="btn btn-secondary" onclick="backToOrders()" style="margin-bottom:12px;">← К списку</button>
        <div class="order-detail-header">
          <div class="num" id="detailNum">—</div>
          <div class="meta" id="detailMeta"></div>
        </div>
        <table>
          <thead><tr><th>Товар</th><th style="text-align:right;">Кол-во</th></tr></thead>
          <tbody id="detailTable"></tbody>
        </table>
        <div id="detailComment" style="font-size:13px;color:#555;margin-top:8px;"></div>
        <div class="buttons" style="margin-top:16px;">
          <button class="btn btn-excel" onclick="downloadCurrentDetailOrder()">⬇ Скачать Excel</button>
        </div>
      </div>
    </div><!-- /content -->
  </div><!-- /container -->
</div><!-- /app -->

<script>
/* ════════════════════════════════════════════════════════
   STATE
════════════════════════════════════════════════════════ */
let tg = null;
let cart = [];
let lastOrderId = null;
let currentDetailOrderId = null;
let foundProducts = [];   // глобальный массив результатов поиска

/* ════════════════════════════════════════════════════════
   INIT
════════════════════════════════════════════════════════ */
function initApp(){
  const check = () => {
    if(typeof Telegram!=='undefined' && Telegram.WebApp && Telegram.WebApp.initData && Telegram.WebApp.initData.length>5){
      tg = Telegram.WebApp; tg.ready(); tg.expand();
      document.querySelector('.app').style.display='block';
      document.getElementById('error-screen').style.display='none';
      setTodayFilter();
      return true;
    }
    return false;
  };
  if(!check()) setTimeout(check,600);
}
window.onload = initApp;

function setTodayFilter(){
  const now = new Date();
  const y = now.getFullYear();
  const m = String(now.getMonth()+1).padStart(2,'0');
  const d = String(now.getDate()).padStart(2,'0');
  const today = `${y}-${m}-${d}`;
  // default: from beginning of month to today
  document.getElementById('filterFrom').value = `${y}-${m}-01`;
  document.getElementById('filterTo').value = today;
}

/* ════════════════════════════════════════════════════════
   NAVIGATION
════════════════════════════════════════════════════════ */
let activeTab = 'search';

function switchTab(tab){
  activeTab = tab;
  document.querySelectorAll('.nav-btn').forEach((b,i)=>{
    b.classList.toggle('active', ['search','orders'][i]===tab);
  });
  if(tab==='search')  showPage('search');
  if(tab==='orders'){ showPage('orders'); renderOrdersList(); }
}

function showPage(page){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.getElementById(page+'Page').classList.add('active');
  const titles = {
    search:'Форма предзаказа', cart:'Корзина', preview:'Предварительный просмотр',
    success:'Предзаказ создан', orders:'Мои предзаказы',
    orderDetail:'Детали предзаказа'
  };
  document.getElementById('headerTitle').textContent = titles[page]||'';
}

function backToSearch(){
  switchTab('search');
}
function backToOrders(){
  showPage('orders');
  renderOrdersList();
}

/* ════════════════════════════════════════════════════════
   SEARCH
════════════════════════════════════════════════════════ */
async function searchProducts(){
  const query = document.getElementById('searchInput').value.trim();
  if(!query) return showMessage('Введите запрос','error');

  showMessage('⏳ Поиск...','info');
  try{
    const res = await fetch('/api/search',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({query})
    });
    const data = await res.json();
    const list = document.getElementById('productsList');
    if(data.products && data.products.length>0){
      foundProducts = data.products;
      list.innerHTML = data.products.map((p,idx)=>
        `<div class="product" data-idx="${idx}">
           <div class="product-name">${p.name}</div>
         </div>`
      ).join('');
      list.onclick = function(e){
        const card = e.target.closest('.product');
        if(!card) return;
        const i = parseInt(card.dataset.idx);
        const p = foundProducts[i];
        if(p) addToCart(p.name, p.unit||'шт');
      };
      list.style.display='flex';
      showMessage(`✅ Найдено ${data.products.length} товаров`,'success');
    } else {
      list.style.display='none';
      showMessage('❌ Товары не найдены','error');
    }
  } catch(e){
    showMessage('❌ Ошибка соединения','error');
  }
}

/* ════════════════════════════════════════════════════════
   CART
════════════════════════════════════════════════════════ */
function addToCart(name, unit){
  const ex = cart.find(i=>i.name===name);
  if(ex){ ex.qty+=1; }
  else { cart.push({name, qty:1, unit}); }
  showMessage('✅ Добавлено в корзину','success');
  updateCartDisplay();
  showPage('cart');
}

function updateCartDisplay(){
  document.getElementById('cartBody').innerHTML = cart.map((item,i)=>`
    <tr>
      <td style="font-size:12px;">${item.name}</td>
      <td style="text-align:center;">${item.qty} ${item.unit}</td>
      <td style="text-align:center;"><button class="action-btn edit-btn" onclick="editItem(${i})">✎</button></td>
      <td style="text-align:center;"><button class="action-btn remove-btn" onclick="removeItem(${i})">✕</button></td>
    </tr>`).join('');
}

function editItem(i){
  const v = prompt(`Количество для:\n${cart[i].name}`, cart[i].qty);
  if(v!==null && !isNaN(v) && parseInt(v)>0){ cart[i].qty=parseInt(v); updateCartDisplay(); }
}
function removeItem(i){
  if(confirm('Удалить товар?')){ cart.splice(i,1); updateCartDisplay(); }
}
function clearCart(){
  if(confirm('Очистить корзину?')){ cart=[]; updateCartDisplay(); }
}

/* ════════════════════════════════════════════════════════
   PREVIEW
════════════════════════════════════════════════════════ */
function showPreview(){
  if(!cart.length) return showMessage('Корзина пуста','error');
  document.getElementById('previewTable').innerHTML = cart.map(i=>
    `<tr><td style="font-size:12px;">${i.name}</td><td style="text-align:right;">${i.qty} ${i.unit}</td></tr>`
  ).join('');
  document.getElementById('previewCount').textContent = cart.length;
  showPage('preview');
}

/* ════════════════════════════════════════════════════════
   CONFIRM & SEND
════════════════════════════════════════════════════════ */
async function confirmAndSend(){
  if(!cart.length) return;
  const comment = document.getElementById('orderComment').value.trim();
  const timestamp = new Date().toLocaleString('ru-RU');

  // Сохранить на сервере
  const res = await fetch('/api/orders',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({items: cart, comment, timestamp})
  });
  const data = await res.json();
  if(data.error) return showMessage('❌ '+data.error,'error');

  lastOrderId = data.order_id;

  // Показать страницу успеха
  document.getElementById('successOrderNum').textContent = data.number;
  document.getElementById('successOrderDate').textContent = data.created_at;
  showPage('success');

  // Отправить в бот
  if(tg && tg.sendData){
    tg.sendData(JSON.stringify({
      order_number: data.number,
      items: cart,
      comment,
      timestamp
    }));
  }

  cart = [];
  updateCartDisplay();
  document.getElementById('orderComment').value='';
}

function viewLastOrder(){
  if(lastOrderId) openOrderDetail(lastOrderId);
}
function downloadLastOrder(){
  if(lastOrderId) downloadOrder(lastOrderId);
}
function newOrder(){
  switchTab('search');
}

/* ════════════════════════════════════════════════════════
   ORDERS LIST
════════════════════════════════════════════════════════ */
async function renderOrdersList(){
  const from = document.getElementById('filterFrom').value;
  const to   = document.getElementById('filterTo').value;

  const res = await fetch(`/api/orders?from=${from}&to=${to}`);
  const data = await res.json();
  const list = document.getElementById('ordersList');

  if(!data.orders || !data.orders.length){
    list.innerHTML = `<div class="empty-state"><div class="icon">📭</div><p>Предзаказов за период не найдено</p></div>`;
    return;
  }

  list.innerHTML = data.orders.map(o=>`
    <div class="order-card" onclick="openOrderDetail(${o.id})">
      <div class="order-card-header">
        <span class="order-card-num">${o.number}</span>
        <span class="order-card-date">${o.created_at}</span>
      </div>
      <div class="order-card-count">📦 ${o.item_count} поз. ${o.comment?'💬':''}</div>
    </div>
  `).join('');
}

async function openOrderDetail(orderId){
  currentDetailOrderId = orderId;
  const res = await fetch(`/api/orders/${orderId}`);
  const o = await res.json();
  if(o.error) return showMessage('❌ '+o.error,'error');

  document.getElementById('detailNum').textContent = o.number;
  document.getElementById('detailMeta').textContent = o.created_at + ' · ' + o.item_count + ' позиций';
  document.getElementById('detailTable').innerHTML = o.items.map(i=>
    `<tr><td style="font-size:12px;">${i.name}</td><td style="text-align:right;">${i.qty} ${i.unit}</td></tr>`
  ).join('');
  document.getElementById('detailComment').textContent = o.comment ? '💬 '+o.comment : '';
  showPage('orderDetail');
}

function downloadCurrentDetailOrder(){
  if(currentDetailOrderId) downloadOrder(currentDetailOrderId);
}

function downloadOrder(orderId){
  window.location.href = `/api/orders/${orderId}/excel`;
}

/* ════════════════════════════════════════════════════════
   UTILS
════════════════════════════════════════════════════════ */
function showMessage(text, type){
  const msg = document.getElementById('message');
  msg.textContent = text;
  msg.className = `message ${type}`;
  clearTimeout(msg._t);
  msg._t = setTimeout(()=>msg.className='message', 4000);
}
</script>
</body>
</html>'''

# ============================================================================
# API ROUTES
# ============================================================================

@app.route('/webapp')
def webapp():
    return render_template_string(MINI_APP_HTML)

# ── Product search (proxy to local API) ─────────────────────────────────────
@app.route('/api/search', methods=['POST'])
def search():
    try:
        data = request.get_json() or {}
        query = data.get('query','').strip()
        if not query:
            return jsonify({"error":"Пустой запрос","products":[]})
        response = session.post(
            'https://criteria-waviness-entangled.ngrok-free.dev/api/search',
            json={'query': query}, timeout=15, verify=False
        )
        return jsonify(response.json())
    except Exception as e:
        print(f"Proxy error: {e}")
        return jsonify({"error":"Локальный сервер недоступен","products":[]})

# ── Create order ─────────────────────────────────────────────────────────────
@app.route('/api/orders', methods=['POST'])
def create_order():
    try:
        data = request.get_json() or {}
        items   = data.get('items', [])
        comment = data.get('comment','')
        if not items:
            return jsonify({"error":"Пустой заказ"})
        oid, num = next_order_id()
        now = datetime.now().strftime("%d.%m.%Y %H:%M")
        ORDERS[oid] = {
            "id": oid,
            "number": num,
            "created_at": now,
            "items": items,
            "comment": comment,
            "item_count": len(items)
        }
        return jsonify({"order_id": oid, "number": num, "created_at": now})
    except Exception as e:
        return jsonify({"error": str(e)})

# ── List orders ───────────────────────────────────────────────────────────────
@app.route('/api/orders', methods=['GET'])
def list_orders():
    from_str = request.args.get('from','')
    to_str   = request.args.get('to','')
    try:
        dt_from = datetime.strptime(from_str, "%Y-%m-%d").date() if from_str else date.min
        dt_to   = datetime.strptime(to_str,   "%Y-%m-%d").date() if to_str   else date.max
    except:
        dt_from, dt_to = date.min, date.max

    result = []
    for o in ORDERS.values():
        try:
            order_date = datetime.strptime(o['created_at'], "%d.%m.%Y %H:%M").date()
        except:
            order_date = date.today()
        if dt_from <= order_date <= dt_to:
            result.append(o)

    result.sort(key=lambda x: x['id'], reverse=True)
    return jsonify({"orders": result})

# ── Get single order ──────────────────────────────────────────────────────────
@app.route('/api/orders/<int:order_id>', methods=['GET'])
def get_order(order_id):
    o = ORDERS.get(order_id)
    if not o:
        return jsonify({"error":"Заказ не найден"})
    return jsonify(o)

# ── Download order as Excel (шаблон акта ТТЗ) ────────────────────────────────
@app.route('/api/orders/<int:order_id>/excel', methods=['GET'])
def download_order_excel(order_id):
    o = ORDERS.get(order_id)
    if not o:
        return jsonify({"error":"Заказ не найден"}), 404
    if not EXCEL_AVAILABLE:
        return jsonify({"error":"openpyxl не установлен"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Акт"

    # ── стили ──
    def tn(s='thin'): return Side(style=s)
    def no():         return Side(style=None)
    def ab(): return Border(left=tn(), right=tn(), top=tn(), bottom=tn())
    def lno(): return Border(left=tn(), top=tn(), bottom=tn())

    f10b = Font(name='Times New Roman', size=10, bold=True)
    f9b  = Font(name='Times New Roman', size=9,  bold=True)
    f9   = Font(name='Times New Roman', size=9)
    f8   = Font(name='Times New Roman', size=8)
    cen  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rgt  = Alignment(horizontal='right',  vertical='center')

    # ── ширины колонок (11 колонок, как в оригинале) ──
    widths = [10.5, 10.5, 10.5, 10.5, 10.5, 10.5, 11.5, 11.3, 14.5, 10.5, 17.3]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    def M(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    def W(r,c,val,fnt=None,aln=None,brd=None):
        cell = ws.cell(row=r, column=c, value=val)
        if fnt: cell.font = fnt
        if aln: cell.alignment = aln
        if brd: cell.border = brd
        return cell

    # ── дата на русском ──
    months = ['января','февраля','марта','апреля','мая','июня',
              'июля','августа','сентября','октября','ноября','декабря']
    now = datetime.now()
    date_ru = f"{now.day} {months[now.month-1]} {now.year} г."

    # ── СТРОКИ 2-4: Приложение ──
    M(2,9,2,11); W(2,9,"Приложения №01", f9, rgt)
    M(3,9,3,11); W(3,9,"к договору комисси №712", f9, rgt)
    M(4,9,4,11); W(4,9,"7 февраля 2022 г.", f9, rgt)

    # ── СТРОКА 6: Заголовок ──
    ws.row_dimensions[6].height = 22.5
    M(6,1,6,11); W(6,1, f"АКТ  приема-передачи товара № {o['number']}", f10b, cen)

    # ── СТРОКА 8: Город / дата ──
    ws.row_dimensions[8].height = 22.5
    M(8,1,8,2);   W(8,1,"г. Ташкент", f9, lft)
    M(8,10,8,11); W(8,10, date_ru, f9, rgt)

    # ── СТРОКИ 10-13: Вводный текст ──
    ws.row_dimensions[10].height = 60
    M(10,1,13,11)
    W(10,1,
      'СП ООО "Ташкентский трубный завод имени В.Л. Гальперина", в лице директора Суцепиной В. Д., '
      'действующего на основании Устава, именуемое в дальнейшем Комитент, с одной стороны и '
      '"МЧЖ ARMOR HAND", в лице директора Исаханова С.С, действующего на основании Устава, '
      ' именуемое в дальнейшем Комиссионер, с другой стороны, составили настоящий Акт о нижеследующем:',
      f9, lft)

    # ── СТРОКИ 14-15: п.1 ──
    M(14,2,15,11)
    W(14,2,
      '1. В соответствии с п.3.2 Договора комиссии №712 от 7 февраля 2022 г.. Комитент передает, '
      'а Комиссионер принимает Товар следующего ассортимента и количества:',
      f9, lft)

    # ── СТРОКИ 16-17: Шапка таблицы ──
    ws.row_dimensions[16].height = 27.75
    M(16,1,17,1);   W(16,1,"№ п/п", f9b, cen, ab())
    M(16,2,17,5);   W(16,2,"Наименование", f9b, cen, ab())
    M(16,6,17,6);   W(16,6,"Ед. изм", f9b, cen, ab())
    M(16,7,17,7);   W(16,7,"Кол-во", f9b, cen, ab())
    M(16,8,17,8);   W(16,8,"Цена без НДС, сум", f9b, cen, ab())
    M(16,9,17,9);   W(16,9,"Цена, включая НДС, сум", f9b, cen, ab())
    M(16,10,17,11); W(16,10,"Сумма, включая НДС сум", f9b, cen, ab())

    # ── ТОВАРНЫЕ СТРОКИ ──
    items = o['items']
    cur = 18
    for idx, item in enumerate(items, 1):
        r1, r2 = cur, cur+1
        ws.row_dimensions[r1].height = 16.875
        ws.row_dimensions[r2].height = 16.875
        M(r1,1,r2,1);   W(r1,1, idx,                    f8, cen, ab())
        M(r1,2,r2,5);   W(r1,2, item.get('name',''),    f8, lft, ab())
        M(r1,6,r2,6);   W(r1,6, item.get('unit','шт'),  f8, cen, ab())
        M(r1,7,r2,7);   W(r1,7, item.get('qty',1),      f8, cen, ab())
        M(r1,8,r2,8);   W(r1,8, "",                      f8, cen, ab())
        M(r1,9,r2,9);   W(r1,9, "",                      f8, cen, ab())
        M(r1,10,r2,11); W(r1,10,"",                      f8, cen, ab())
        cur += 2

    # ── ИТОГО ──
    ws.row_dimensions[cur].height = 19.125
    M(cur,1,cur,5)
    W(cur,1,"Итого:", f9b, rgt,
      Border(left=tn(),top=tn(),bottom=tn(),right=no()))
    M(cur,6,cur,9)
    W(cur,6,"",f9, cen,
      Border(left=no(),right=no(),top=tn(),bottom=tn()))
    M(cur,10,cur,11)
    W(cur,10,"", f9b, rgt, ab())

    # ── п.2 ──
    r = cur + 1
    M(r,2,r+2,11)
    W(r,2,
      '2.Принятый Комиссионером товар обладает качеством и ассортиментом, соответствующим требованиям Договора.'
      'Товара поставлен в установленные в Договоре сроки. Комиссионер не имеет никаких претензий к принятому товару.',
      f9, lft)
    r += 3

    # ── п.3 ──
    M(r,2,r+1,11)
    W(r,2,
      '3.Настоящий Акт составлен в двух экземплярах, имеющих равную юридическую силу, '
      'по одному экземпляру для каждой из Сторон и является неотъемлемой частью Договора между Сторонами.',
      f9, lft)
    r += 3

    # ── ПОДПИСИ ──
    M(r,2,r,5);   W(r,2,"КОМИТЕНТ", f9b, cen)
    M(r,7,r,11);  W(r,7,"КОМИССИОНЕР", f9b, cen)
    r += 2
    M(r,2,r,5);   W(r,2,"Директор       Суцепина В. Д.", f9, lft)
    M(r,7,r,11);  W(r,7,"Директор       Исаханов С.С", f9, lft)
    r += 2
    M(r,2,r,5);   W(r,2,"Ст. спец. реализации", f9, lft)
    r += 2
    M(r,2,r,5);   W(r,2,"Товар отпустил:", f9, lft)
    M(r,7,r,11);  W(r,7,"Товар принял: Yakubxodjayev U", f9, lft)
    r += 2
    M(r,2,r,5);   W(r,2,"М.П.", f9, cen)
    M(r,7,r,11);  W(r,7,"М.П.", f9, cen)

    # ── печать ──
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{o['number'].replace('/','-')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============================================================================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("\n🛡️ ARMOR HAND Cloud v7.0")
    print("  ✅ Выгрузка Excel строго по шаблону акта ТТЗ")
    print("  ✅ Предзаказы с фильтром по периоду")
    app.run(host='0.0.0.0', port=port, debug=False)
